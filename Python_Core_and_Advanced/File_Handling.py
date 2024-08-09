#Types of Files
"""
1. Text files
2. Binary files
3. CSV files
4. Excel files

Opening file :

Syntax : f=open(filename,mode)

Modes
w= writing data to the file, it will override the exiting data inside a file
r= read data from the exiting file,it will not override
a= Open an exiting file for append mode operation.it won't Override the exiting data
r+= To read and write data into the file
w+= To write and read the data . It will override existing data
a+ =To append and read data from the file. It won't Override the exiting data.

For Binary files : rb,wb,ab,r+b,w+b,a+b


Closing ==>close()
f.close()

Various Properties of a file Object

name ===> Name to Open the particular file
Mode ==>Mode in which the file is opened
closed ==>Returns the boolean Value it indicates file is closed or Opened.
readable() ==> Returns the boolean value
writeable() ==>Returns the boolean value

"""
#1. Text files
f=open("abc.txt",'w')
print("The File Name is :",f.name)
print("The File Mode is :",f.mode)
print("is File Readable :",f.readable())
print("is File Writable :",f.writable())
print("is File Opened : ",f.closed)
f.close()
print("is File Opened : ",f.closed)

#Write the data
#Ex 1
f=open("abc.txt",'w')
f.write("QACircle\n")
f.write("Software\n")
f.write("Training\n")
f.write("Academy\n")
print("The data is written to the file successfully")
f.close()

#Ex 2
# f=open("abc.txt",'w')
# list=["Amogh\n","Bhavya\n","Kavyashree\n","Rohith\n","Hema\n"]
# f.writelines(list)
# f.close()

#Read the data from the file
"""
read() ==> To read the data from the file 
read(n) ===> To read 'n' characters from the file
readline() ==> To read only one line
readlines() ==> TO read all the lines into list.
"""
#Ex 1:
f=open("abc.txt",'r')
data=f.read()
print(data)
f.close()

#Ex 2:
f=open("abc.txt",'r')
data=f.read(10)
print(data)
f.close()

#Ex 3:
f=open("abc.txt",'r')
line1=f.readline()
print(line1, end='')
line2=f.readline()
print(line2, end='')
line3=f.readline()
print(line3, end='')
f.close()
print("****************************")
#Ex 4:
f=open("abc.txt",'r')
lines=f.readlines()
for line in lines:
    print(line,end="")
f.close()

#With Statement:
with open("abc.txt",'w') as f:
    f.write("QACircle\n")
    f.write("Software\n")
    f.write("Training\n")
    f.write("Academy\n")
    print("is File Closed :",f.closed)
print("is File Closed :",f.closed)

# Handling Binary file
# f1=open("C:\\Users\\dell\\Downloads\\XYZ.jpeg",'rb')
# f2=open("NewQAC_image.jpg",'wb')
# bytes=f1.read()
# print(bytes)
# f2.write(bytes)

# #Handling CSV file :
# #Writing the Data to the CSV file
# """
# CSV ==> Comma Separated Value
# """
# print("******CSV ==> Comma Separated Value********")
# import csv
# with open("emp.csv",'w',newline="") as f:
#     w=csv.writer(f) #Retuen the CSV writer Object
#     w.writerow(["ENO","ENAME","ESAL","EADDR"])
#     n=int(input("Enter the number of Employees"))
#     for i in range(n):
#         eno=input("Enter the Employee No")
#         ename=input("Enter the Employee Name")
#         esal=float(input("Enter the Employee Salary"))
#         eaddr=input("Enter the Employee address ")
#         w.writerow([eno,ename,esal,eaddr])
#
# print("Total Employee data is written to CSV file successfully")
#
# #Read data from CSV file
# import csv
# f=open("emp.csv",'r')
# r=csv.reader(f) #Return the CSV reader Object
# data=list(r)
# print(data)
# for line in data:
#     for word in line:
#         print(word,"\t",end='')
#     print()

# Excel File
"""
Case 1 : How to read data from the excel file
Case 2 : How to write the data into excel file
"""
# Case 1 : How to read data from the Excel file
import openpyxl

file="D:\\QC20_Python_SDET\\QACircle_Read.xlsx"

workbook=openpyxl.load_workbook(file)

sheet=workbook["Sheet1"]

#Count the number of rows in this particular Excel sheet
rows=sheet.max_row
print(rows)

#Count the number of columns in this particular Excel sheet
column=sheet.max_column
print(column)

#Read all the rows and columns present in the excel sheet
for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end='  ')
    print()

# Case 2 : How to write the data into Excel file
import openpyxl

file="D:\\QC20_Python_SDET\\QACircle_Write.xlsx"

workbook=openpyxl.load_workbook(file)

sheet=workbook.active

#Count the number of rows in this particular Excel sheet
rows=sheet.max_row
print(rows)

#Count the number of columns in this particular Excel sheet
column=sheet.max_column
print(column)

#Write all the rows and columns present in the excel sheet
# for r in range(1,rows+1):
#     for c in range(1,column+1):
#         sheet.cell(r,c).value="Hello"
#     print()

sheet.cell(2,1).value="Basavana Gouda"
sheet.cell(2,2).value="Python"

sheet.cell(3,1).value="Chaya"
sheet.cell(3,2).value="SQL"


workbook.save(file)


#Case 3
import openpyxl
wb=openpyxl.Workbook()
sheet=wb.active

c1=sheet.cell(row=1,column=1)
c1.value="Hi Good Morning"

c2=sheet.cell(row=1,column=2)
c2.value="Welcome to the Class"

c3=sheet['A3']
c3.value="Selenium"

wb.save("sample.xlsx")
